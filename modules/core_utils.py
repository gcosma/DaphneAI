# ===============================================
# FILE: modules/core_utils.py (COMPLETE VERSION)
# ===============================================

import re
import logging
import unicodedata
import hashlib
import base64
from datetime import datetime
from typing import Dict, List, Any, Optional, Union, Tuple
from pathlib import Path
import pandas as pd
import urllib3
from sklearn.feature_extraction.text import TfidfVectorizer
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import math
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import time
import os
import zipfile
import asyncio

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s: %(message)s",
    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global headers for all requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Referer": "https://judiciary.uk/",
}

# ===============================================
# TEXT CLEANING AND PROCESSING
# ===============================================

def clean_text(text: str) -> str:
    """Clean text while preserving structure and metadata formatting"""
    if not text:
        return ""

    try:
        text = str(text)
        text = unicodedata.normalize("NFKD", text)

        # Character replacements for common encoding issues
        replacements = {
            "â€™": "'", "â€œ": '"', "â€": '"', "â€¦": "...", 'â€"': "-", "â€¢": "•",
            "Â": " ", "\u200b": "", "\uf0b7": "", "\u2019": "'", "\u201c": '"',
            "\u201d": '"', "\u2013": "-", "\u2022": "•",
        }

        for encoded, replacement in replacements.items():
            text = text.replace(encoded, replacement)

        # Remove HTML tags
        text = re.sub(r"<[^>]+>", "", text)
        
        # Keep only printable characters and newlines
        text = "".join(
            char if char.isprintable() or char == "\n" else " " for char in text
        )
        
        # Normalize whitespace
        text = re.sub(r" +", " ", text)
        text = re.sub(r"\n+", "\n", text)

        return text.strip()

    except Exception as text_error:
        logging.error(f"Error in clean_text: {text_error}")
        return ""

def clean_text_for_modeling(text: str) -> str:
    """Clean text for BERT modeling and machine learning"""
    if not text:
        return ""

    try:
        text = clean_text(text)
        text = text.lower()

        # Remove dates, times, and reference numbers
        text = re.sub(r"\b\d{4}[-/]\d{2}[-/]\d{2}\b", "", text)
        text = re.sub(r"\b\d{1,2}:\d{2}\b", "", text)
        text = re.sub(r"\bref\w*\s*[-:\s]?\s*\w+[-\d]+\b", "", text, flags=re.IGNORECASE)
        
        # Remove legal terms that don't add semantic value
        legal_terms = r"\b(coroner|inquest|hearing|evidence|witness|statement|report|dated|signed)\b"
        text = re.sub(legal_terms, "", text, flags=re.IGNORECASE)

        # Keep only letters and spaces
        text = re.sub(r"[^a-z\s]", " ", text)
        text = re.sub(r"\s+", " ", text)
        
        # Remove short words (less than 3 characters)
        text = " ".join(word for word in text.split() if len(word) > 2)

        return text.strip() if len(text.split()) >= 3 else ""

    except Exception as modeling_error:
        logging.error(f"Error in clean_text_for_modeling: {modeling_error}")
        return ""

def extract_concern_text(content: str) -> str:
    """
    Enhanced: Extract complete concern text from PDF report content with robust section handling
    
    This version handles government documents and inquiry reports with improved pattern matching.
    """
    if pd.isna(content) or not isinstance(content, str):
        return ""

    try:
        # Enhanced concern section identifiers for various document types
        concern_patterns = [
            # Standard coroner patterns
            r"CORONER'S\s+CONCERNS?:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            r"MATTERS?\s+OF\s+CONCERN:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            r"The\s+MATTERS?\s+OF\s+CONCERN:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            
            # Variations with "are" after the identifier
            r"CORONER'S\s+CONCERNS?\s+are:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            r"MATTERS?\s+OF\s+CONCERN\s+are:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            r"CONCERNS?\s+IDENTIFIED:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            
            # Handle OCR issues where spaces are missing
            r"TheMATTERS?\s*OF\s*CONCERN:?\s*(.*?)(?=ACTION\s+SHOULD\s+BE\s+TAKEN|CONCLUSIONS|YOUR\s+RESPONSE|COPIES|SIGNED:|DATED\s+THIS|NEXT\s+STEPS|YOU\s+ARE\s+UNDER\s+A\s+DUTY|$)",
            
            # Government inquiry patterns (generic for any inquiry)
            r"(?i)(?:key\s+)?concerns?:?\s*(.*?)(?=recommendations?|next\s+steps|conclusions?|$)",
            r"(?i)matters?\s+of\s+concern:?\s*(.*?)(?=recommendations?|next\s+steps|conclusions?|$)",
            r"(?i)issues?\s+identified:?\s*(.*?)(?=recommendations?|next\s+steps|conclusions?|$)",
        ]
        
        # Enhanced text normalization to handle PDF extraction issues
        content_norm = _normalize_pdf_text(content)
        
        # Try each pattern and keep the longest/best match
        best_concerns_text = ""
        best_confidence = 0.0
        
        for pattern in concern_patterns:
            try:
                matches = re.finditer(pattern, content_norm, re.IGNORECASE | re.DOTALL | re.MULTILINE)
                
                for match in matches:
                    extracted_text = match.group(1).strip()
                    
                    if extracted_text:
                        # Calculate confidence score for this extraction
                        confidence = _calculate_extraction_confidence(extracted_text, pattern)
                        
                        # Keep the best extraction based on length and confidence
                        if (len(extracted_text) > len(best_concerns_text) and confidence >= best_confidence) or confidence > best_confidence + 0.1:
                            best_concerns_text = extracted_text
                            best_confidence = confidence
                            
            except re.error as regex_error:
                logging.warning(f"Regex error with pattern: {pattern[:50]}... Error: {regex_error}")
                continue
        
        # Post-process the extracted text
        if best_concerns_text:
            return _post_process_concerns_text(best_concerns_text)
        
        return ""
        
    except Exception as extract_error:
        logging.error(f"Error in extract_concern_text: {extract_error}")
        return ""

def _normalize_pdf_text(text: str) -> str:
    """Normalize PDF text to handle common extraction issues"""
    if not text:
        return ""
    
    # Basic normalization - collapse whitespace but preserve structure
    text = ' '.join(text.split())
    
    # Fix common OCR issues
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)  # Add spaces between words
    text = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', text)  # Add spaces before numbers
    text = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', text)  # Add spaces after numbers
    
    # Normalize punctuation spacing
    text = re.sub(r'\s*:\s*', ': ', text)
    text = re.sub(r'\s*\.\s*', '. ', text)
    
    return text

def _calculate_extraction_confidence(text: str, pattern: str) -> float:
    """Calculate confidence score for extracted text"""
    confidence = 0.5  # Base confidence
    
    # Length bonus (longer extractions generally more complete)
    if len(text) > 100:
        confidence += 0.2
    if len(text) > 300:
        confidence += 0.1
    
    # Content quality indicators
    if any(word in text.lower() for word in ['concern', 'risk', 'death', 'future', 'safety']):
        confidence += 0.1
    
    # Structure indicators (numbered points, paragraphs)
    if re.search(r'\d+\.', text):
        confidence += 0.1
    
    return min(confidence, 1.0)

def _post_process_concerns_text(text: str) -> str:
    """Clean and post-process extracted concerns text"""
    if not text:
        return ""
    
    # Remove leading/trailing whitespace
    text = text.strip()
    
    # Clean up excessive whitespace
    text = re.sub(r'\s+', ' ', text)
    
    # Try to end at a complete sentence if text seems cut off
    if not text.endswith(('.', '!', '?', ':')):
        # Find the last complete sentence
        last_period = text.rfind('.')
        last_exclamation = text.rfind('!')
        last_question = text.rfind('?')
        
        last_punctuation = max(last_period, last_exclamation, last_question)
        
        # If we found punctuation and it's in the latter part of the text, cut there
        if last_punctuation != -1 and last_punctuation > len(text) * 0.7:
            text = text[:last_punctuation + 1].strip()
    
    # Remove any leftover artifacts
    text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
    text = re.sub(r'^[:\-\s]+', '', text)  # Remove leading punctuation
    
    return text.strip()

# ===============================================
# SECURITY AND VALIDATION
# ===============================================

class SecurityValidator:
    """Security validation for file operations and user input"""
    
    @staticmethod
    def validate_file_upload(file_content: bytes, filename: str) -> bool:
        """Validate uploaded files for security"""
        MAX_SIZE = 500 * 1024 * 1024  # 500MB limit
        
        if len(file_content) > MAX_SIZE:
            raise ValueError(f"File too large: {len(file_content)/1024/1024:.1f}MB > {MAX_SIZE/1024/1024}MB")
        
        allowed_extensions = {'.pdf', '.txt', '.docx', '.xlsx', '.csv', '.json'}
        file_ext = Path(filename).suffix.lower()
        if file_ext not in allowed_extensions:
            raise ValueError(f"Invalid file type: {file_ext}")
        
        # Additional validation for PDF files
        if file_ext == '.pdf':
            if not file_content.startswith(b'%PDF'):
                raise ValueError("Invalid PDF file format")
        
        return True
    
    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """Sanitize filename for safe storage"""
        filename = os.path.basename(filename)
        filename = re.sub(r'[^\w\-_\.]', '_', filename)
        
        # Ensure filename isn't too long
        if len(filename) > 255:
            name, ext = os.path.splitext(filename)
            filename = name[:255-len(ext)] + ext
        
        return filename
    
    @staticmethod
    def validate_text_input(text: str, max_length: int = 10000) -> str:
        """Sanitize and validate text input"""
        if not text:
            return ""
        
        # Convert to string and limit length
        text = str(text)[:max_length]
        
        # Remove potentially harmful content
        text = re.sub(r'<script.*?</script>', '', text, flags=re.IGNORECASE | re.DOTALL)
        text = re.sub(r'javascript:', '', text, flags=re.IGNORECASE)
        text = re.sub(r'on\w+\s*=', '', text, flags=re.IGNORECASE)
        
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text)
        
        return text.strip()
    
    @staticmethod
    def validate_search_query(query: str) -> Tuple[bool, str]:
        """Validate search query for safety"""
        if not query or not query.strip():
            return False, "Search query cannot be empty"
        
        query = query.strip()
        
        if len(query) < 3:
            return False, "Search query must be at least 3 characters"
        
        if len(query) > 1000:
            return False, "Search query too long (max 1000 characters)"
        
        # Check for potential injection attempts
        suspicious_patterns = [
            r'union\s+select', r'drop\s+table', r'delete\s+from',
            r'insert\s+into', r'update\s+.*set'
        ]
        
        for pattern in suspicious_patterns:
            if re.search(pattern, query, re.IGNORECASE):
                return False, "Query contains suspicious content"
        
        return True, "Valid search query"
    
    @staticmethod
    def get_file_hash(file_content: bytes) -> str:
        """Generate hash for file content"""
        return hashlib.sha256(file_content).hexdigest()

# ===============================================
# METADATA EXTRACTION
# ===============================================

def extract_metadata(content: str) -> Dict[str, Any]:
    """Extract metadata from document content"""
    metadata = {
        "extraction_date": datetime.now().isoformat(),
        "content_length": len(content),
        "word_count": len(content.split()) if content else 0
    }
    
    if not content:
        return metadata
    
    try:
        # Name patterns for deceased person (coroner reports)
        name_patterns = [
            r"(?:In the matter of|Concerning the death of|The death of)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)",
            r"([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:died|\s+was)"
        ]
        
        for pattern in name_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                metadata["deceased_name"] = match.group(1).strip()
                break
        
        # Date patterns
        date_patterns = [
            r"Date\s+of\s+report:?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
            r"Dated\s+this\s+(\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})"
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                metadata["date_of_report"] = match.group(1).strip()
                break
        
        # Government document patterns
        if re.search(r'(?i)government\s+response', content):
            metadata["document_type"] = "government_response"
        elif re.search(r'(?i)inquiry\s+report', content):
            metadata["document_type"] = "inquiry_report"
        elif re.search(r'(?i)cabinet\s+office', content):
            metadata["document_type"] = "cabinet_office_document"
        
        return metadata
        
    except Exception as metadata_error:
        logging.error(f"Error extracting metadata: {metadata_error}")
        return metadata

def extract_government_document_metadata(content: str) -> Dict[str, Any]:
    """Extract metadata specific to government documents"""
    metadata = extract_metadata(content)
    
    try:
        # Government-specific patterns
        gov_patterns = {
            'department': r'(?i)(?:department\s+for|ministry\s+of|cabinet\s+office)\s+([^.\n]+)',
            'command_paper': r'(?i)(?:cp|cm|command\s+paper)\s+(\d+)',
            'isbn': r'(?i)isbn\s+([\d\-]+)',
            'publication_date': r'(?i)(?:published|may|dated)\s+(\d{4})',
            'minister': r'(?i)(?:minister\s+for|secretary\s+of\s+state\s+for)\s+([^.\n]+)'
        }
        
        for key, pattern in gov_patterns.items():
            match = re.search(pattern, content)
            if match:
                metadata[key] = match.group(1).strip()
        
        # Check for inquiry types (generic patterns for any UK inquiry)
        inquiry_patterns = {
            'inquiry_type': r'(?i)(\w+(?:\s+\w+)*)\s+inquiry',
            'commission_type': r'(?i)(\w+(?:\s+\w+)*)\s+commission',
            'investigation_type': r'(?i)(\w+(?:\s+\w+)*)\s+investigation',
            'review_type': r'(?i)(\w+(?:\s+\w+)*)\s+review'
        }
        
        for key, pattern in inquiry_patterns.items():
            match = re.search(pattern, content)
            if match:
                inquiry_name = match.group(1).strip()
                # Clean up common prefixes
                inquiry_name = re.sub(r'(?i)^(?:the\s+|public\s+|independent\s+)', '', inquiry_name)
                metadata[key] = inquiry_name
                break
        
        # Check document classification
        if re.search(r'(?i)government\s+response', content):
            metadata['document_classification'] = 'government_response'
        elif re.search(r'(?i)(?:inquiry\s+report|final\s+report)', content):
            metadata['document_classification'] = 'inquiry_report'
        elif re.search(r'(?i)(?:interim\s+report|progress\s+report)', content):
            metadata['document_classification'] = 'interim_report'
        elif re.search(r'(?i)terms\s+of\s+reference', content):
            metadata['document_classification'] = 'terms_of_reference'
        
        return metadata
        
    except Exception as gov_metadata_error:
        logging.error(f"Error extracting government metadata: {gov_metadata_error}")
        return metadata

# ===============================================
# DATA PROCESSING AND VALIDATION
# ===============================================

def process_scraped_data(df: pd.DataFrame) -> pd.DataFrame:
    """Process scraped data with cleaning and validation"""
    if df is None or len(df) == 0:
        return df
    
    try:
        # Clean text fields
        if 'Content' in df.columns:
            df['Content'] = df['Content'].apply(lambda x: clean_text(str(x)) if pd.notna(x) else "")
        
        if 'Title' in df.columns:
            df['Title'] = df['Title'].apply(lambda x: clean_text(str(x)) if pd.notna(x) else "")
        
        # Extract concerns if not already present
        if 'Extracted_Concerns' not in df.columns and 'Content' in df.columns:
            df['Extracted_Concerns'] = df['Content'].apply(extract_concern_text)
        
        # Process dates
        if 'date_of_report' in df.columns:
            df['date_of_report'] = pd.to_datetime(df['date_of_report'], errors='coerce')
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove duplicates if ref column exists
        if 'ref' in df.columns:
            df = df.drop_duplicates(subset=['ref'], keep='first')
        
        return df
        
    except Exception as processing_error:
        logging.error(f"Error processing scraped data: {processing_error}")
        return df

def is_response_document(row: pd.Series) -> bool:
    """Check if a document is a response document"""
    try:
        for i in range(1, 10):  # Check PDF_1 to PDF_9
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                return True
        return False
    except Exception as check_error:
        logging.error(f"Error checking response document: {check_error}")
        return False

def validate_data(df: pd.DataFrame) -> Dict[str, Any]:
    """Validate DataFrame and return quality metrics"""
    validation_result = {
        "is_valid": True,
        "row_count": len(df),
        "column_count": len(df.columns),
        "issues": [],
        "warnings": [],
        "data_quality_score": 0.0
    }
    
    if df.empty:
        validation_result["is_valid"] = False
        validation_result["issues"].append("DataFrame is empty")
        return validation_result
    
    try:
        # Check for required columns
        required_columns = ['Title', 'Content']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            validation_result["warnings"].append(f"Missing recommended columns: {missing_columns}")
        
        # Check for empty content
        quality_scores = []
        if 'Content' in df.columns:
            empty_content = df['Content'].isna().sum() + (df['Content'] == "").sum()
            if empty_content > 0:
                validation_result["warnings"].append(f"{empty_content} rows have empty content")
            
            # Content completeness score
            content_completeness = (len(df) - empty_content) / len(df)
            quality_scores.append(content_completeness)
            if content_completeness < 0.8:
                validation_result["warnings"].append(f"Content completeness: {content_completeness:.1%}")
        
        # Date completeness
        if "date_of_report" in df.columns:
            date_completeness = df["date_of_report"].notna().mean()
            quality_scores.append(date_completeness)
            if date_completeness < 0.7:
                validation_result["warnings"].append(f"Date completeness: {date_completeness:.1%}")
        
        # Calculate overall quality score
        if quality_scores:
            validation_result["data_quality_score"] = sum(quality_scores) / len(quality_scores)
        
        # Check data quality
        duplicate_count = df.duplicated().sum()
        if duplicate_count > 0:
            validation_result["warnings"].append(f"{duplicate_count} duplicate rows found")
        
        # Check for duplicates based on Title and ref if available
        if "Title" in df.columns and "ref" in df.columns:
            duplicate_count = df.duplicated(subset=["Title", "ref"]).sum()
            if duplicate_count > 0:
                validation_result["warnings"].append(f"Found {duplicate_count} potential duplicates")
        
        # Check data types
        if "date_of_report" in df.columns:
            if not pd.api.types.is_datetime64_any_dtype(df["date_of_report"]):
                validation_result["warnings"].append("Date column is not in datetime format")
        
        return validation_result
        
    except Exception as validation_error:
        validation_result["issues"].append(f"Validation error: {validation_error}")
        validation_result["is_valid"] = False
        logging.error(f"Data validation error: {validation_error}")
        return validation_result

# ===============================================
# DOCUMENT STRUCTURE ANALYSIS
# ===============================================

def detect_inquiry_document_structure(content: str) -> Dict[str, Any]:
    """Detect the structure and type of inquiry document"""
    structure_info = {
        'has_recommendations': False,
        'has_responses': False,
        'has_table_of_contents': False,
        'has_numbered_sections': False,
        'document_structure': 'unknown',
        'estimated_recommendations': 0,
        'estimated_responses': 0
    }
    
    try:
        content_lower = content.lower()
        
        # Check for recommendations
        rec_indicators = [
            r'(?i)\brecommendations?\b',
            r'(?i)\brecommend\s+that\b',
            r'(?i)\bshould\s+(?:be|ensure|implement)\b',
            r'(?i)\bmust\s+(?:be|ensure|implement)\b'
        ]
        
        for pattern in rec_indicators:
            if re.search(pattern, content):
                structure_info['has_recommendations'] = True
                break
        
        # Count potential recommendations
        numbered_recs = len(re.findall(r'(?i)(?:^|\n)\s*(?:recommendation\s+)?\d+[\.\)]\s', content))
        structure_info['estimated_recommendations'] = numbered_recs
        
        # Check for responses
        resp_indicators = [
            r'(?i)\bresponse\s+to\b',
            r'(?i)\bgovernment\s+(?:accepts?|rejects?|agrees?)\b',
            r'(?i)\b(?:accepted|rejected|partially\s+accepted)\b',
            r'(?i)\bimplementation\s+(?:plan|strategy)\b'
        ]
        
        for pattern in resp_indicators:
            if re.search(pattern, content):
                structure_info['has_responses'] = True
                break
        
        # Count potential responses
        numbered_resps = len(re.findall(r'(?i)(?:^|\n)\s*(?:response\s+to\s+)?(?:recommendation\s+)?\d+[:\.\)]\s', content))
        structure_info['estimated_responses'] = numbered_resps
        
        # Check for table of contents
        toc_indicators = [
            r'(?i)(?:table\s+of\s+)?contents',
            r'(?i)list\s+of\s+chapters',
            r'(?i)\d+\.\d+\s+[A-Z]'  # Pattern like "1.5 Recommendations"
        ]
        
        for pattern in toc_indicators:
            if re.search(pattern, content):
                structure_info['has_table_of_contents'] = True
                break
        
        # Check for numbered sections
        numbered_sections = len(re.findall(r'(?:^|\n)\s*\d+\.\d+\s+[A-Z]', content))
        if numbered_sections >= 3:
            structure_info['has_numbered_sections'] = True
        
        # Determine document structure
        if structure_info['has_recommendations'] and structure_info['has_responses']:
            structure_info['document_structure'] = 'combined_inquiry_and_response'
        elif structure_info['has_recommendations']:
            structure_info['document_structure'] = 'inquiry_report'
        elif structure_info['has_responses']:
            structure_info['document_structure'] = 'government_response'
        elif structure_info['has_table_of_contents']:
            structure_info['document_structure'] = 'structured_document'
        
        return structure_info
        
    except Exception as structure_error:
        logging.error(f"Error detecting document structure: {structure_error}")
        return structure_info

def extract_inquiry_themes(content: str) -> List[str]:
    """Extract common themes from inquiry documents"""
    themes = []
    
    try:
        content_lower = content.lower()
        
        # Common inquiry themes (generic patterns for any inquiry)
        theme_patterns = {
            'safety': r'(?i)\b(?:safety|safe|unsafe|hazard|risk|danger)\b',
            'communication': r'(?i)\b(?:communication|inform|notif|alert|warn)\b',
            'training': r'(?i)\b(?:training|education|competenc|skill|qualif)\b',
            'procedure': r'(?i)\b(?:procedure|process|protocol|guideline|standard)\b',
            'oversight': r'(?i)\b(?:oversight|supervis|monitor|audit|review)\b',
            'resource': r'(?i)\b(?:resource|funding|staff|budget|capacity)\b',
            'governance': r'(?i)\b(?:governance|management|leadership|accountab)\b',
            'technology': r'(?i)\b(?:technology|system|equipment|software|digital)\b',
            'culture': r'(?i)\b(?:culture|attitude|behavior|mindset)\b',
            'transparency': r'(?i)\b(?:transparen|open|disclosure|public)\b'
        }
        
        for theme, pattern in theme_patterns.items():
            if len(re.findall(pattern, content)) >= 3:  # At least 3 mentions
                themes.append(theme)
        
        return themes
        
    except Exception as theme_error:
        logging.error(f"Error extracting themes: {theme_error}")
        return themes

# ===============================================
# UTILITY FUNCTIONS
# ===============================================

def format_date_uk(date_input: Union[str, datetime]) -> str:
    """Format date in UK format"""
    if pd.isna(date_input):
        return ""
    
    try:
        if isinstance(date_input, str):
            # Try to parse string date
            date_obj = pd.to_datetime(date_input)
        else:
            date_obj = date_input
        
        return date_obj.strftime("%d/%m/%Y")
    except Exception as date_error:
        logging.error(f"Error formatting date: {date_error}")
        return str(date_input)

def create_document_identifier(title: str, date: str = None) -> str:
    """Create a unique identifier for a document"""
    try:
        # Clean title for identifier
        clean_title = re.sub(r'[^\w\s]', '', title)
        clean_title = re.sub(r'\s+', '_', clean_title)
        
        # Add date if provided
        if date:
            clean_date = re.sub(r'[^\d]', '', str(date))
            return f"{clean_title}_{clean_date}"
        
        return clean_title
        
    except Exception as id_error:
        logging.error(f"Error creating document identifier: {id_error}")
        return "unknown_document"

def deduplicate_documents(data: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate documents while preserving unique entries"""
    try:
        if data is None or len(data) == 0:
            return data
        
        # Create document identifiers for rows that have the required columns
        if 'Title' in data.columns:
            data["doc_id"] = data.apply(lambda row: create_document_identifier(
                str(row.get("Title", "")), 
                str(row.get("ref", ""))
            ), axis=1)
            
            # Remove duplicates based on identifier
            data_deduped = data.drop_duplicates(subset=["doc_id"], keep="first")
            
            # Remove the temporary identifier column
            data_deduped = data_deduped.drop(columns=["doc_id"])
            
            logging.info(f"Deduplicated {len(data)} -> {len(data_deduped)} documents")
            return data_deduped
        
        return data
        
    except Exception as dedup_error:
        logging.error(f"Error in deduplication: {dedup_error}")
        return data

def combine_document_text(row: pd.Series) -> str:
    """Combine all text content from a document row for analysis"""
    text_parts = []
    
    try:
        if pd.notna(row.get("Title")): 
            text_parts.append(str(row["Title"]))
        if pd.notna(row.get("Content")): 
            text_parts.append(str(row["Content"]))
        
        # Iterate through potential PDF content columns
        for i in range(1, 10):  # Assuming up to PDF_9_Content
            pdf_col_name = f"PDF_{i}_Content"
            if pdf_col_name in row.index and pd.notna(row.get(pdf_col_name)):
                text_parts.append(str(row[pdf_col_name]))
                
        return " ".join(text_parts)
        
    except Exception as combine_error:
        logging.error(f"Error combining document text: {combine_error}")
        return ""

def filter_by_document_types(df: pd.DataFrame, doc_types: List[str]) -> pd.DataFrame:
    """Filter DataFrame by document types (Report/Response)"""
    if not doc_types: 
        return df
    
    try:
        is_response_series = df.apply(is_response_document, axis=1)
        
        if "Report" in doc_types and "Response" not in doc_types:
            return df[~is_response_series]
        elif "Response" in doc_types and "Report" not in doc_types:
            return df[is_response_series]
        # If both are selected, return all
        return df
        
    except Exception as filter_error:
        logging.error(f"Error filtering by document types: {filter_error}")
        return df

def filter_by_categories(df: pd.DataFrame, categories: List[str]) -> pd.DataFrame:
    """Filter DataFrame by categories if category column exists"""
    try:
        if not categories or 'Category' not in df.columns:
            return df
        
        return df[df['Category'].isin(categories)]
        
    except Exception as category_error:
        logging.error(f"Error filtering by categories: {category_error}")
        return df

def perform_advanced_keyword_search(df: pd.DataFrame, keywords: List[str], search_columns: List[str] = None) -> pd.DataFrame:
    """Perform advanced keyword search across specified columns"""
    if not keywords or df.empty:
        return df
    
    if search_columns is None:
        search_columns = ['Title', 'Content', 'Extracted_Concerns']
    
    # Available columns in dataframe
    available_columns = [col for col in search_columns if col in df.columns]
    
    if not available_columns:
        return df
    
    try:
        # Create search mask
        mask = pd.Series([False] * len(df))
        
        for keyword in keywords:
            keyword_lower = keyword.lower().strip()
            if not keyword_lower:
                continue
                
            # Search in each column
            for col in available_columns:
                col_mask = df[col].astype(str).str.lower().str.contains(
                    keyword_lower, case=False, na=False, regex=False
                )
                mask = mask | col_mask
        
        return df[mask]
        
    except Exception as search_error:
        logging.error(f"Error in keyword search: {search_error}")
        return df

# ===============================================
# EXPORT FUNCTIONS
# ===============================================

def export_to_excel(df: pd.DataFrame, filename: str = "export.xlsx") -> bytes:
    """Export DataFrame to Excel with formatting"""
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return output.getvalue()
        
    except Exception as export_error:
        logging.error(f"Error exporting to Excel: {export_error}")
        return b""

def analyze_document_collection(documents: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Analyze a collection of documents and provide insights"""
    analysis = {
        'total_documents': len(documents),
        'document_types': {},
        'themes': {},
        'date_range': {},
        'quality_metrics': {}
    }
    
    try:
        if not documents:
            return analysis
        
        # Analyze document types
        type_counts = {}
        theme_counts = {}
        dates = []
        
        for doc in documents:
            # Document type analysis
            doc_type = doc.get('document_type', 'unknown')
            type_counts[doc_type] = type_counts.get(doc_type, 0) + 1
            
            # Theme analysis
            doc_themes = doc.get('themes', [])
            for theme in doc_themes:
                theme_counts[theme] = theme_counts.get(theme, 0) + 1
            
            # Date analysis
            doc_date = doc.get('date_of_report')
            if doc_date:
                try:
                    dates.append(pd.to_datetime(doc_date))
                except:
                    pass
        
        analysis['document_types'] = type_counts
        analysis['themes'] = theme_counts
        
        # Date range analysis
        if dates:
            analysis['date_range'] = {
                'earliest': min(dates).strftime('%Y-%m-%d'),
                'latest': max(dates).strftime('%Y-%m-%d'),
                'span_years': (max(dates) - min(dates)).days / 365.25
            }
        
        return analysis
        
    except Exception as analysis_error:
        logging.error(f"Error analyzing document collection: {analysis_error}")
        return analysis

def generate_insights_report(analysis: Dict[str, Any]) -> str:
    """Generate a text report from document analysis"""
    try:
        report_lines = [
            "# Document Collection Analysis Report",
            f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "",
            "## Overview",
            f"Total Documents: {analysis.get('total_documents', 0)}",
            ""
        ]
        
        # Document types section
        doc_types = analysis.get('document_types', {})
        if doc_types:
            report_lines.extend([
                "## Document Types",
                ""
            ])
            for doc_type, count in sorted(doc_types.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / analysis['total_documents']) * 100
                report_lines.append(f"- {doc_type.title()}: {count} ({percentage:.1f}%)")
            report_lines.append("")
        
        # Themes section
        themes = analysis.get('themes', {})
        if themes:
            report_lines.extend([
                "## Common Themes",
                ""
            ])
            for theme, count in sorted(themes.items(), key=lambda x: x[1], reverse=True)[:10]:
                report_lines.append(f"- {theme.title()}: {count} documents")
            report_lines.append("")
        
        # Date range section
        date_range = analysis.get('date_range', {})
        if date_range:
            report_lines.extend([
                "## Date Range",
                f"- Earliest Document: {date_range.get('earliest', 'Unknown')}",
                f"- Latest Document: {date_range.get('latest', 'Unknown')}",
                f"- Time Span: {date_range.get('span_years', 0):.1f} years",
                ""
            ])
        
        return "\n".join(report_lines)
        
    except Exception as report_error:
        logging.error(f"Error generating insights report: {report_error}")
        return "Error generating report"

# ===============================================
# STREAMLIT HELPERS
# ===============================================

def create_download_link(data: bytes, filename: str, link_text: str) -> str:
    """Create a download link for Streamlit"""
    try:
        b64_data = base64.b64encode(data).decode()
        return f'<a href="data:application/octet-stream;base64,{b64_data}" download="{filename}">{link_text}</a>'
    except Exception as link_error:
        logging.error(f"Error creating download link: {link_error}")
        return "Download unavailable"

def display_progress_bar(current: int, total: int, text: str = "Processing"):
    """Display a progress bar in Streamlit"""
    try:
        if total > 0:
            progress = current / total
            st.progress(progress, text=f"{text}: {current}/{total} ({progress:.1%})")
        else:
            st.info("No items to process")
    except Exception as progress_error:
        logging.error(f"Error displaying progress: {progress_error}")

# ===============================================
# ADVANCED SEARCH AND FILTERING
# ===============================================

def create_text_search_index(df: pd.DataFrame, text_columns: List[str] = None) -> Dict[str, Any]:
    """Create a text search index for faster searching"""
    if text_columns is None:
        text_columns = ['Title', 'Content', 'Extracted_Concerns']
    
    try:
        # Combine text from specified columns
        combined_text = []
        available_columns = [col for col in text_columns if col in df.columns]
        
        for idx, row in df.iterrows():
            text_parts = []
            for col in available_columns:
                if pd.notna(row[col]):
                    text_parts.append(str(row[col]))
            combined_text.append(" ".join(text_parts))
        
        # Create TF-IDF vectorizer
        vectorizer = TfidfVectorizer(
            max_features=5000,
            stop_words='english',
            ngram_range=(1, 2),
            min_df=1,
            max_df=0.95
        )
        
        # Fit and transform
        tfidf_matrix = vectorizer.fit_transform(combined_text)
        
        return {
            'vectorizer': vectorizer,
            'matrix': tfidf_matrix,
            'feature_names': vectorizer.get_feature_names_out()
        }
        
    except Exception as index_error:
        logging.error(f"Error creating search index: {index_error}")
        return {}

def search_with_index(search_index: Dict[str, Any], query: str, top_k: int = 10) -> List[int]:
    """Search using the text index and return document indices"""
    try:
        if not search_index or not query:
            return []
        
        vectorizer = search_index['vectorizer']
        tfidf_matrix = search_index['matrix']
        
        # Transform query
        query_vector = vectorizer.transform([query])
        
        # Calculate similarities
        from sklearn.metrics.pairwise import cosine_similarity
        similarities = cosine_similarity(query_vector, tfidf_matrix).flatten()
        
        # Get top results
        top_indices = similarities.argsort()[-top_k:][::-1]
        
        # Filter out results with very low similarity
        threshold = 0.01
        relevant_indices = [idx for idx in top_indices if similarities[idx] > threshold]
        
        return relevant_indices
        
    except Exception as search_error:
        logging.error(f"Error searching with index: {search_error}")
        return []

# ===============================================
# MODULE EXPORTS
# ===============================================

__all__ = [
    'clean_text',
    'clean_text_for_modeling', 
    'extract_concern_text',
    'SecurityValidator',
    'extract_metadata',
    'extract_government_document_metadata',
    'process_scraped_data',
    'is_response_document',
    'validate_data',
    'detect_inquiry_document_structure',
    'extract_inquiry_themes',
    'format_date_uk',
    'create_document_identifier',
    'deduplicate_documents',
    'combine_document_text',
    'filter_by_document_types',
    'filter_by_categories',
    'perform_advanced_keyword_search',
    'export_to_excel',
    'analyze_document_collection',
    'generate_insights_report',
    'create_download_link',
    'display_progress_bar',
    'create_text_search_index',
    'search_with_index',
    '_normalize_pdf_text',
    '_calculate_extraction_confidence',
    '_post_process_concerns_text'
]
